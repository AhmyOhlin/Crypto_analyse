import json, os
from datetime import datetime
from coinmarketcap import Market
import pandas as pd
filename_xls = 'XXXXXXX.xlsx'
xls = pd.ExcelFile(filename_xls)
df = xls.parse(xls.sheet_names[0])
#print(df)
import openpyxl
print(os.getcwd())

wb = openpyxl.load_workbook('2018_05_cc.xlsx')
wb2 = openpyxl.load_workbook('test.xlsx')
sheet = wb.get_sheet_by_name('Trade_Jan18')
sheet_2 = wb2.get_sheet_by_name('Feuil1')

coinmarketcap = Market()
filename = 'Eth.json'

# get data from api and copare it with my_list to store the specified entities in JSON-File
data = coinmarketcap.ticker(start=0, limit=1400, convert='EUR')
my_list =['EOS','NEO','NEM',...]

def store_data_to_json(): #  data of your coins will be stored in JSON-file
    global json_data
    try:  # does the data structure exist yet? Let's try opening the file...
        with open(filename) as feedjson:
            json_data = json.load(feedjson)
    except FileNotFoundError:  # this must be the first execution. Create an empty data structure.
        json_data = {"my_coins": []}
    for entity in data:
        #print(entity)
        if entity['name'] in my_list:
            json_data['my_coins'].append(entity)

    # overwrite the old json dict with the updated one
    with open(filename, "w") as feedjson:
        json.dump(json_data, feedjson, indent=4)
        feedjson.write('\n')

store_data_to_json()

def data_analysis():
    global liste
    liste = []
    for item in json_data['my_coins']:
        if item['name'] in  my_list:
            liste.append({item['name']: float(item['price_eur'])})

data_analysis()


dicto = {}
for element in liste:
    for key , val in element.items():
        dicto[key] = val
sorted_list = (sorted(zip(dicto.keys(),dicto.values())))
print('sorted_list_from_CMC',sorted_list)

cmc_price_list = []
for i in sorted_list:
    cmc_price_list.append(i[1])
print('cmc_price_list',cmc_price_list)

#get old data from excel file and merge it with the new data from Coinmarketcap API
rows = 20
columns = 12
liste = []
print(len(cmc_price_list))
for i in range(1, rows + 1):
    liste.append([])
for r in range(1, rows + 1):
    for c in range(1, columns + 1):
        if c == 6:
            e = sheet.cell(row=r, column=c)
            b = liste[r-1].append(cmc_price_list[r-1])
        else:
            e = sheet.cell(row=r, column=c)
            liste[r - 1].append(e.value)


# write the new data to excel file
for r in range(1, rows+1):
    print('r', r,c)
    for c in range(1,columns+1):
        j = sheet_2.cell(row=r, column=c)
        j.value= liste[r-1][c-1]      
wb.save('test.xlsx')


# print the stored entities of excel file
for _ in liste:
    print(_)

